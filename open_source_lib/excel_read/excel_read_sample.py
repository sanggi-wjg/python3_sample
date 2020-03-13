"""
Required openpyxl
    pip install openpyxl
document 가 잘되어 있고 속도가 비교적 좋다고 한다.

https://openpyxl.readthedocs.io/en/stable/tutorial.html#accessing-many-cells
"""
import pprint

from openpyxl import load_workbook

if __name__ == '__main__':
    workbook = load_workbook(filename = 'sample_read.xlsx', read_only = True, data_only = True)
    print(workbook)

    try:
        sheets = workbook.sheetnames
        targetSheetName = sheets[1]
        # print(sheets)
        # print(targetSheetName)

        worksheets = workbook.worksheets
        targetSheet = workbook.worksheets[1]
        # print(worksheets)
        # print(targetSheet)

        readData = { }

        for rowNo, row in enumerate(targetSheet.iter_rows()):
            if rowNo == 0: continue
            invoiceNo, where, weight, cost = '', '', '', ''

            for cellNo, cell in enumerate(row):
                if cellNo == 2:
                    invoiceNo = str(cell.value).strip()

                elif cellNo == 4:
                    where = str(cell.value)

                elif cellNo == 5:
                    weight = str(float(cell.value))

                elif cellNo == 6:
                    cost = str(float(cell.value))

            if not invoiceNo: continue

            readData.setdefault(invoiceNo, {
                'invoiceNo': invoiceNo,
                'where'    : where,
                'weight'   : weight,
                'cost'     : cost,
            })

        readData = dict(sorted(readData.items(), key = lambda x: x[0]))
        # print(readData)
        pp = pprint.PrettyPrinter(2)
        pp.pprint(readData)

    except Exception as e:
        print(e.__str__())

    finally:
        workbook.close()
